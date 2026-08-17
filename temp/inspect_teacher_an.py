import sys
import openpyxl
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

excel_path = Path(r"C:\Users\Love\Documents\Codex\MD\tonggv0517082026.xlsx")
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb.active

teacher_cols = {}
for col in range(4, ws.max_column + 1):
    t_name = str(ws.cell(row=2, column=col).value or "").strip()
    if t_name:
        teacher_cols[t_name] = col

print("\n=== CLASS 8A4 SCHEDULE ON THU 4 SANG (Row 23 to 27) ===")
for p in range(5):
    r = 23 + p
    subjects_in_8a4 = []
    for t_name, col in teacher_cols.items():
        val = str(ws.cell(row=r, column=col).value or "").strip()
        if "8A4" in val or "8/4" in val:
            subjects_in_8a4.append(f"{val} ({t_name})")
    print(f"  Tiết {p+1}: {', '.join(subjects_in_8a4) if subjects_in_8a4 else 'OFF/Empty'}")

print("\n=== CLASS 9A11 SCHEDULE ON THU 4 SANG (Row 23 to 27) ===")
for p in range(5):
    r = 23 + p
    subjects_in_9a11 = []
    for t_name, col in teacher_cols.items():
        val = str(ws.cell(row=r, column=col).value or "").strip()
        if "9A11" in val or "9/11" in val:
            subjects_in_9a11.append(f"{val} ({t_name})")
    print(f"  Tiết {p+1}: {', '.join(subjects_in_9a11) if subjects_in_9a11 else 'OFF/Empty'}")
