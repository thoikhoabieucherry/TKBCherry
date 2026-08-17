import sys
import os
import openpyxl
import json
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')

def parse_schedule_structure(ws):
    # Rows structure
    schedule_grid = {} # (day, period) -> row_index
    row_info = {} # row_index -> (day, period)
    
    current_day = ""
    for r in range(6, ws.max_row + 1):
        day_val = ws.cell(r, 1).value
        period_val = ws.cell(r, 2).value
        if day_val is not None and str(day_val).strip() != "":
            current_day = str(day_val).strip()
        
        p_str = str(period_val).strip() if period_val is not None else ""
        row_info[r] = (current_day, p_str)
    
    # Columns structure for classes and teachers
    col_info = {} # col_index -> (name, session) (e.g. ('6/1', 'Sáng'))
    header_name_row = 4
    header_sess_row = 5
    
    current_name = ""
    for c in range(3, ws.max_column + 1):
        name_val = ws.cell(header_name_row, c).value
        sess_val = ws.cell(header_sess_row, c).value
        if name_val is not None and str(name_val).strip() != "":
            current_name = str(name_val).strip()
        sess_str = str(sess_val).strip() if sess_val is not None else ""
        col_info[c] = (current_name, sess_str)
        
    return row_info, col_info

def load_schedule_matrix(filepath, sheet_name='TKB_LOP_SC'):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name]
    row_info, col_info = parse_schedule_structure(ws)
    
    # matrix: (entity_name, session, day, period) -> cell_value
    grid = {}
    for r, (day, period) in row_info.items():
        if not period or not day:
            continue
        for c, (entity, sess) in col_info.items():
            val = ws.cell(r, c).value
            cell_val = str(val).strip() if val is not None else ""
            grid[(entity, sess, day, period)] = cell_val
            
    return grid, row_info, col_info

def analyze_gaps(grid, is_teacher=True):
    # Analyze teaching gaps for each teacher/day/session
    # A day/session has periods 1, 2, 3, 4, 5
    # Find all occupied periods
    # Gap is empty periods between min_period and max_period
    gaps_by_entity = defaultdict(list)
    # entity -> list of {day, session, teaching_periods, gap_lengths, gap_periods}
    
    # Get all entities, days, sessions
    entities = sorted(list(set(k[0] for k in grid.keys())))
    sessions = sorted(list(set(k[1] for k in grid.keys())))
    days = ["Hai", "Ba", "Tư", "Năm", "Sáu", "Bảy"] # approximate or dynamic
    
    # Let's dynamically get days and periods
    all_days = []
    for k in grid.keys():
        if k[2] not in all_days:
            all_days.append(k[2])
            
    periods_by_sess = defaultdict(set)
    for k in grid.keys():
        periods_by_sess[k[1]].add(k[3])
        
    for entity in entities:
        for sess in sessions:
            for day in all_days:
                # get teaching periods
                teaching = []
                day_periods = []
                for p in range(1, 6):
                    p_str = str(p)
                    val = grid.get((entity, sess, day, p_str), "")
                    if (entity, sess, day, p_str) in grid:
                        day_periods.append(p)
                        if val != "":
                            teaching.append((p, val))
                
                if len(teaching) >= 2:
                    t_indices = [t[0] for t in teaching]
                    min_p = min(t_indices)
                    max_p = max(t_indices)
                    
                    # check intermediate periods
                    current_gap = 0
                    current_gap_start = 0
                    for p in range(min_p, max_p + 1):
                        val = grid.get((entity, sess, day, str(p)), "")
                        if val == "":
                            if current_gap == 0:
                                current_gap_start = p
                            current_gap += 1
                        else:
                            if current_gap > 0:
                                gaps_by_entity[entity].append({
                                    'day': day,
                                    'sess': sess,
                                    'gap_length': current_gap,
                                    'gap_periods': list(range(current_gap_start, current_gap_start + current_gap)),
                                    'teaching': teaching
                                })
                                current_gap = 0
                    if current_gap > 0:
                        gaps_by_entity[entity].append({
                            'day': day,
                            'sess': sess,
                            'gap_length': current_gap,
                            'gap_periods': list(range(current_gap_start, current_gap_start + current_gap)),
                            'teaching': teaching
                        })
    return gaps_by_entity

def diff_grids(grid1, grid2):
    diffs = []
    all_keys = set(grid1.keys()).union(set(grid2.keys()))
    for k in all_keys:
        v1 = grid1.get(k, "")
        v2 = grid2.get(k, "")
        if v1 != v2:
            diffs.append({
                'entity': k[0],
                'session': k[1],
                'day': k[2],
                'period': k[3],
                'before': v1,
                'after': v2
            })
    return diffs

if __name__ == "__main__":
    dir_vesion = r"C:\Users\Love\Documents\Codex\temp\vesion"
    base_file = os.path.join(dir_vesion, "base.xlsx")
    
    # Load base GV
    grid_base_gv, _, _ = load_schedule_matrix(base_file, 'TKB_GV_SC')
    grid_base_lop, _, _ = load_schedule_matrix(base_file, 'TKB_LOP_SC')
    
    gaps_base = analyze_gaps(grid_base_gv)
    print("=== BASE GAPS SUMMARY (TKB_GV_SC) ===")
    total_2_gaps = 0
    total_1_gaps = 0
    total_3_gaps = 0
    for entity, glist in gaps_base.items():
        g2 = [g for g in glist if g['gap_length'] == 2]
        g1 = [g for g in glist if g['gap_length'] == 1]
        g3 = [g for g in glist if g['gap_length'] >= 3]
        if g2 or g1 or g3:
            print(f"Teacher {entity}: {len(g1)} gap(s) of 1, {len(g2)} gap(s) of 2, {len(g3)} gap(s) >= 3")
            for g in g2:
                print(f"   -> GAP 2: Day {g['day']}, Sess {g['sess']}, Periods {g['gap_periods']}, Teaching: {g['teaching']}")
        total_2_gaps += len(g2)
        total_1_gaps += len(g1)
        total_3_gaps += len(g3)
        
    print(f"\nTOTAL: {total_1_gaps} 1-period gaps, {total_2_gaps} 2-period gaps, {total_3_gaps} >=3-period gaps\n")
    
    # Compare with each teacher version in vesion/
    for fname in sorted(os.listdir(dir_vesion)):
        if fname == "base.xlsx" or not fname.endswith(".xlsx"):
            continue
        t_file = os.path.join(dir_vesion, fname)
        t_grid_gv, _, _ = load_schedule_matrix(t_file, 'TKB_GV_SC')
        t_grid_lop, _, _ = load_schedule_matrix(t_file, 'TKB_LOP_SC')
        
        diffs_gv = diff_grids(grid_base_gv, t_grid_gv)
        diffs_lop = diff_grids(grid_base_lop, t_grid_lop)
        t_gaps = analyze_gaps(t_grid_gv)
        
        teacher_name = fname.replace(".xlsx", "")
        print(f"==================================================")
        print(f"OPTIMIZATION FOR: {fname} (Target teacher might be {teacher_name})")
        print(f"Total cell differences in TKB_GV: {len(diffs_gv)}, in TKB_LOP: {len(diffs_lop)}")
        
        # Check gaps of the target teacher before and after
        base_t_gaps = gaps_base.get(teacher_name, [])
        new_t_gaps = t_gaps.get(teacher_name, [])
        print(f"Before gaps for {teacher_name}: {[ (g['day'], g['sess'], g['gap_periods']) for g in base_t_gaps ]}")
        print(f"After gaps for {teacher_name}:  {[ (g['day'], g['sess'], g['gap_periods']) for g in new_t_gaps ]}")
        
        print("\nDIFF DETAILS in TKB_LOP (What changed in classes):")
        # Group diffs by class
        diff_by_class = defaultdict(list)
        for d in diffs_lop:
            diff_by_class[d['entity']].append(d)
        for cls, dlist in diff_by_class.items():
            print(f"  Class {cls}:")
            for d in dlist:
                print(f"    Day {d['day']} {d['session']} P{d['period']}: '{d['before']}' -> '{d['after']}'")
                
        print("\nDIFF DETAILS in TKB_GV (What changed for teachers):")
        diff_by_gv = defaultdict(list)
        for d in diffs_gv:
            diff_by_gv[d['entity']].append(d)
        for gv, dlist in diff_by_gv.items():
            print(f"  Teacher {gv}:")
            for d in dlist:
                print(f"    Day {d['day']} {d['session']} P{d['period']}: '{d['before']}' -> '{d['after']}'")
        print("\n")
