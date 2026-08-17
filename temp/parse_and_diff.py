import sys
import os
import openpyxl
import json

sys.stdout.reconfigure(encoding='utf-8')

def parse_tkb_sheet(ws):
    data = []
    for r in range(1, ws.max_row + 1):
        row_vals = []
        for c in range(1, ws.max_column + 1):
            val = ws.cell(r, c).value
            row_vals.append("" if val is None else str(val).strip())
        data.append(row_vals)
    return data

def compare_sheets(sheet1_data, sheet2_data):
    diffs = []
    max_r = max(len(sheet1_data), len(sheet2_data))
    for r in range(max_r):
        r1 = sheet1_data[r] if r < len(sheet1_data) else []
        r2 = sheet2_data[r] if r < len(sheet2_data) else []
        max_c = max(len(r1), len(r2))
        for c in range(max_c):
            v1 = r1[c] if c < len(r1) else ""
            v2 = r2[c] if c < len(r2) else ""
            if v1 != v2:
                diffs.append({
                    "row": r + 1,
                    "col": c + 1,
                    "val_before": v1,
                    "val_after": v2
                })
    return diffs

def analyze_directory(dir_path):
    print(f"=== Analyzing Directory: {dir_path} ===")
    files = [f for f in os.listdir(dir_path) if f.endswith('.xlsx')]
    print(f"Files: {files}")
    
    base_file = os.path.join(dir_path, "base.xlsx")
    if not os.path.exists(base_file):
        print("No base.xlsx found!")
        return
        
    wb_base = openpyxl.load_workbook(base_file, data_only=True)
    base_lop = parse_tkb_sheet(wb_base['TKB_LOP_SC'])
    base_gv = parse_tkb_sheet(wb_base['TKB_GV_SC'])
    
    # Print headers of TKB_LOP_SC and TKB_GV_SC to understand structure
    print("\n--- Structure of TKB_LOP_SC ---")
    for r in range(min(5, len(base_lop))):
        print(f"Row {r+1}: {base_lop[r][:15]}")
        
    print("\n--- Structure of TKB_GV_SC ---")
    for r in range(min(5, len(base_gv))):
        print(f"Row {r+1}: {base_gv[r][:15]}")

if __name__ == "__main__":
    analyze_directory(r"C:\Users\Love\Documents\Codex\temp\vesion")
