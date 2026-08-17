import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

def check_pccm_9_9():
    pccm_file = r"C:\Users\Love\Documents\Codex\phancong\binhtridong\BTĐ PHÂN CÔNG CHUYÊN MÔN  NĂM HỌC 2026 - 2027.xlsx"
    wb = openpyxl.load_workbook(pccm_file, data_only=True)
    print("Sheets in PCCM:", wb.sheetnames)
    
    for sname in wb.sheetnames:
        ws = wb[sname]
        for r in range(1, ws.max_row + 1):
            row_v = [str(ws.cell(r, c).value or "") for c in range(1, ws.max_column + 1)]
            for c_idx, val in enumerate(row_v):
                if "9/9" in val or "9A9" in val:
                    print(f"Sheet {sname} (Row {r}, Col {c_idx+1}): {row_v[:12]}")

if __name__ == "__main__":
    check_pccm_9_9()
