import openpyxl
import os
import glob
import pandas as pd
import json

def inspect_file(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_names = wb.sheetnames
    print(f"File: {filepath}")
    print(f"Sheets: {sheet_names}")
    for name in sheet_names[:3]:
        sheet = wb[name]
        print(f"Sheet {name}: {sheet.max_row} rows, {sheet.max_column} cols")
        for r in range(1, min(10, sheet.max_row + 1)):
            row_vals = [str(sheet.cell(r, c).value) if sheet.cell(r, c).value is not None else "" for c in range(1, min(15, sheet.max_column + 1))]
            if any(row_vals):
                print(f"  Row {r}: {row_vals}")

if __name__ == "__main__":
    vesion_base = r"C:\Users\Love\Documents\Codex\temp\vesion\base.xlsx"
    inspect_file(vesion_base)
