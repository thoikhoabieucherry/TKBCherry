import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

def inspect_sheet1():
    filepath = r"C:\Users\Love\Documents\Codex\MD\tonggv0417082026.xlsx"
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Sheet1']
    print(f"Sheet1 dimensions: max_row={ws.max_row}, max_column={ws.max_column}")
    
    for r in range(1, min(15, ws.max_row + 1)):
        row_vals = [str(ws.cell(r, c).value) if ws.cell(r, c).value is not None else "" for c in range(1, min(20, ws.max_column + 1))]
        print(f"Row {r:2d}: {row_vals}")

if __name__ == "__main__":
    inspect_sheet1()
