import sys
import os
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

def search_for_9a9():
    base_dir = r"C:\Users\Love\Documents\Codex"
    print("Searching for 9A9 in all excel files...")
    
    for root, dirs, files in os.walk(base_dir):
        if "node_modules" in root or ".git" in root or "TKBCherry-backups" in root:
            continue
        for f in files:
            if f.endswith(".xlsx") and not f.startswith("~$"):
                p = os.path.join(root, f)
                try:
                    wb = openpyxl.load_workbook(p, read_only=True)
                    sheet_names = wb.sheetnames
                    wb.close()
                    # Open data_only to check 9A9
                    wb = openpyxl.load_workbook(p, data_only=True)
                    found = False
                    for sname in wb.sheetnames:
                        ws = wb[sname]
                        for r in range(1, min(10, ws.max_row + 1)):
                            for c in range(1, min(150, ws.max_column + 1)):
                                val = str(ws.cell(r, c).value or "")
                                if "9A9" in val or "9/9" in val:
                                    print(f"Found 9A9 in file: {p} (Sheet: {sname}, Cell R{r}C{c}: {val})")
                                    found = True
                                    break
                            if found: break
                        if found: break
                    wb.close()
                except Exception as e:
                    pass

if __name__ == "__main__":
    search_for_9a9()
