import sys
import os
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

def analyze_pccm_btd():
    pccm_file = r"C:\Users\Love\Documents\Codex\phancong\binhtridong\BTĐ PHÂN CÔNG CHUYÊN MÔN  NĂM HỌC 2026 - 2027.xlsx"
    daxep_file = r"C:\Users\Love\Documents\Codex\phancong\tesst\da xep.xlsx"
    nghi_file = r"C:\Users\Love\Documents\Codex\phancong\tesst\nghi.xlsx"
    
    print("=== CHECKING DA XEP.XLSX ===")
    if os.path.exists(daxep_file):
        wb = openpyxl.load_workbook(daxep_file, data_only=True)
        print("Sheets in da xep:", wb.sheetnames)
        # Check 9/9
        ws = wb['TKB_LOP_SC']
        for r in range(4, ws.max_row + 1):
            for c in range(3, ws.max_column + 1):
                header = ws.cell(4, c).value
                if header == "9/9" or header == "9A9":
                    sess = ws.cell(5, c).value
                    lessons = []
                    for row_idx in range(6, ws.max_row + 1):
                        day = ws.cell(row_idx, 1).value
                        p = ws.cell(row_idx, 2).value
                        v = ws.cell(row_idx, c).value
                        if v:
                            lessons.append((day, p, v))
                    print(f"Lớp {header} ({sess}): Đã xếp {len(lessons)} tiết:")
                    for l in lessons:
                        print(f"   Thứ {l[0]} Tiết {l[1]}: {l[2]}")
                        
    print("\n=== CHECKING NGHI.XLSX ===")
    if os.path.exists(nghi_file):
        wb = openpyxl.load_workbook(nghi_file, data_only=True)
        print("Sheets in nghi:", wb.sheetnames)
        for sname in wb.sheetnames:
            ws = wb[sname]
            print(f"Sheet {sname}:")
            for r in range(1, min(10, ws.max_row+1)):
                row_v = [str(ws.cell(r, c).value or "") for c in range(1, min(20, ws.max_column+1))]
                if any(row_v): print(f"  Row {r}: {row_v[:10]}")

if __name__ == "__main__":
    analyze_pccm_btd()
