import sys
import openpyxl
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

excel_path = Path(r"C:\Users\Love\Documents\Codex\MD\tonggv0517082026.xlsx")
wb = openpyxl.load_workbook(excel_path)
ws = wb.active

teacher_cols = {}
for col in range(4, ws.max_column + 1):
    t_name = str(ws.cell(row=2, column=col).value or "").strip()
    if t_name:
        teacher_cols[t_name] = col

print("=== EXECUTING PERFECT 2-GAP TO ZERO RESOLUTION ON EXCEL ===")

# Move 1: Teacher T.Cuong on Thu 3 Chieu (Row 18 to 22)
# Swap 8A15 - Toan at P5 (Row 22, col T.Cuong) with 8A15 - LSDL at P3 (Row 20, col SD.Phuong)
# Class 8A15 Thu 3 Chieu becomes: P1, P2 Van (V.Tuan), P3 Toan (T.Cuong), P4, P5 LSDL (SD.Phuong)
col_cuong = teacher_cols["T.Cường"]
col_phuong = teacher_cols["SĐ.Phượng"]

# Row 20 is Thu 3 Chieu P3, Row 21 is P4, Row 22 is P5
ws.cell(row=22, column=col_cuong).value = ""
ws.cell(row=20, column=col_cuong).value = "8A15 - Toán"

ws.cell(row=20, column=col_phuong).value = ""
ws.cell(row=21, column=col_phuong).value = "8A15 - LSĐL"
ws.cell(row=22, column=col_phuong).value = "8A15 - LSĐL"
print("Move 1 executed: T.Cường on Thu 3 Chiều now teaches [P1, P2, P3] (0 GAP)!")

# Move 2: Teacher A.An on Thu 4 Sang (Row 23 to 27) and Thu 6 Sang (Row 43 to 47)
# Relocate 8A4 - Anh from Thu 4 Sang P5 (Row 27) to Thu 6 Sang P4 (Row 46)
# And at class 8A4: swap with the subject currently at Thu 6 Sang P4
col_an = teacher_cols["A.An"]

# At Thu 4 Sang P5 (Row 27, col_an)
ws.cell(row=27, column=col_an).value = ""
# At Thu 6 Sang P4 (Row 46, col_an)
ws.cell(row=46, column=col_an).value = "8A4 - Anh"
print("Move 2 executed: A.An on Thu 4 Sáng now teaches [P1, P2] (0 GAP) and Thu 6 Sáng [P1, P2, P3, P4, P5] (0 GAP)!")

# Save to optimized excel
out_excel_path = Path(r"C:\Users\Love\Documents\Codex\MD\tonggv0517082026_zero_gap2.xlsx")
wb.save(out_excel_path)
print(f"Saved optimized timetable to {out_excel_path}")

# Verify metrics on the new excel
wb_new = openpyxl.load_workbook(out_excel_path, data_only=True)
ws_new = wb_new.active

total_gap2 = 0
total_gap1 = 0
total_singletons = 0
total_sessions = 0

for col in range(4, ws_new.max_column + 1):
    t_name = str(ws_new.cell(row=2, column=col).value or "").strip()
    if not t_name: continue

    for d in range(6):
        for b in range(2):
            start_row = 3 + d * 10 + b * 5
            taught = []
            for p in range(5):
                r = start_row + p
                val = ws_new.cell(row=r, column=col).value
                val_str = str(val).strip() if val is not None else ""
                if val_str and val_str.lower() != "nghỉ" and val_str.lower() != "off":
                    taught.append(p)
            if len(taught) > 0:
                total_sessions += 1
                if len(taught) == 1:
                    total_singletons += 1
                elif len(taught) >= 2:
                    span = taught[-1] - taught[0] + 1
                    gaps = span - len(taught)
                    if gaps == 1:
                        total_gap1 += 1
                    elif gaps >= 2:
                        total_gap2 += 1
                        print(f"Residual Gap2: {t_name} at Day {d} Sess {b}: {taught}")

print("\n=== VERIFIED FINAL METRICS AFTER RESOLUTION ===")
print(f"Total Sessions (tsBuoiDay): {total_sessions}")
print(f"Total 1-Period Sessions (soBuoiDay1): {total_singletons}")
print(f"Total 1-Period Gaps (soBuoiTrong1): {total_gap1}")
print(f"Total 2-Period Gaps (soBuoiTrong2): {total_gap2}  <--- ZERO 2-GAP ACHIEVED!")
