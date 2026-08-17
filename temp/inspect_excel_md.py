import sys
import openpyxl
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

excel_path = Path(r"C:\Users\Love\Documents\Codex\MD\tonggv0517082026.xlsx")
wb = openpyxl.load_workbook(excel_path, data_only=True)

print("Sheet names in tonggv0517082026.xlsx:", wb.sheetnames)

# Inspect first sheet structure
ws = wb.active
print(f"Active sheet: {ws.title}, max_row: {ws.max_row}, max_column: {ws.max_column}")

for r in range(1, min(30, ws.max_row + 1)):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, min(25, ws.max_column + 1))]
    print(f"Row {r:2d}: {row_vals}")
