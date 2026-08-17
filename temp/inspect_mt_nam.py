import sys
import openpyxl
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

excel_path = Path(r"C:\Users\Love\Documents\Codex\MD\tonggv0517082026.xlsx")
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb.active

teacher_cols = {}
for col in range(4, ws.max_column + 1):
    t_name = str(ws.cell(row=2, column=col).value or "").strip()
    if t_name:
        teacher_cols[t_name] = col

def print_teacher_schedule(t_name):
    col = teacher_cols.get(t_name)
    if not col:
        print(f"Teacher {t_name} not found!")
        return
    print(f"\n=== SCHEDULE FOR {t_name} ===")
    days = ["Thứ 2", "Thứ 3", "Thứ 4", "Thứ 5", "Thứ 6", "Thứ 7"]
    sessions = ["Sáng", "Chiều"]
    for d in range(6):
        for b in range(2):
            start_row = 3 + d * 10 + b * 5
            cells = [str(ws.cell(row=start_row + p, column=col).value or "").strip() for p in range(5)]
            print(f"  {days[d]} {sessions[b]}: {cells}")

print_teacher_schedule("MT.Nam")
print_teacher_schedule("GD.Tâm")
