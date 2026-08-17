from __future__ import annotations

import json
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


def cell_value(value):
    if value is None:
        return None
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            pass
    return value


def inspect(path: Path) -> dict:
    wb = load_workbook(path, data_only=False, read_only=False)
    result = {
        "path": str(path),
        "sheetnames": wb.sheetnames,
        "defined_names": sorted(wb.defined_names),
        "sheets": [],
    }
    for ws in wb.worksheets:
        nonempty = []
        formulas = []
        types = Counter()
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    nonempty.append((cell.coordinate, cell_value(cell.value)))
                    types[cell.data_type] += 1
                    if cell.data_type == "f":
                        formulas.append((cell.coordinate, cell.value))
        result["sheets"].append(
            {
                "title": ws.title,
                "dimensions": ws.calculate_dimension(),
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "nonempty_count": len(nonempty),
                "cell_types": dict(types),
                "formula_count": len(formulas),
                "formulas": formulas[:30],
                "merged_ranges": [str(r) for r in ws.merged_cells.ranges],
                "tables": list(ws.tables),
                "freeze_panes": str(ws.freeze_panes) if ws.freeze_panes else None,
                "sample": nonempty[:120],
            }
        )
    return result


if __name__ == "__main__":
    paths = [Path(p) for p in sys.argv[1:]]
    print(json.dumps([inspect(p) for p in paths], ensure_ascii=False, indent=2))
