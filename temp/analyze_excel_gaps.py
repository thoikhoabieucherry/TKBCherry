import sys
import openpyxl
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

excel_path = Path(r"C:\Users\Love\Documents\Codex\MD\tonggv0517082026.xlsx")
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb.active

# Days and Sessions mapping
# 6 days * 2 sessions * 5 periods = 60 rows (rows 3 to 62)
# Row 2 is header with teacher names

teachers = []
for col in range(4, ws.max_column + 1):
    t_name = ws.cell(row=2, column=col).value
    if t_name and str(t_name).strip():
        teachers.append((col, str(t_name).strip()))

print(f"Total teachers in Excel: {len(teachers)}")

total_gap2 = 0
total_gap1 = 0
total_singletons = 0
total_sessions = 0

gap2_details = []

for col, t_name in teachers:
    t_gap2 = 0
    t_gap1 = 0
    t_singletons = 0
    t_sessions = 0

    # 6 days
    for d in range(6):
        # 2 sessions per day
        for b in range(2):
            start_row = 3 + d * 10 + b * 5
            taught_periods = []
            cells = []
            for p in range(5):
                r = start_row + p
                val = ws.cell(row=r, column=col).value
                val_str = str(val).strip() if val is not None else ""
                cells.append(val_str)
                if val_str and val_str.lower() != "nghỉ" and val_str.lower() != "off":
                    taught_periods.append(p)

            if len(taught_periods) > 0:
                t_sessions += 1
                if len(taught_periods) == 1:
                    t_singletons += 1
                elif len(taught_periods) >= 2:
                    span = taught_periods[-1] - taught_periods[0] + 1
                    gaps = span - len(taught_periods)
                    if gaps == 1:
                        t_gap1 += 1
                    elif gaps >= 2:
                        t_gap2 += 1
                        day_name = ["Thứ 2", "Thứ 3", "Thứ 4", "Thứ 5", "Thứ 6", "Thứ 7"][d]
                        sess_name = ["Sáng", "Chiều"][b]
                        gap2_details.append({
                            "teacher": t_name,
                            "day": day_name,
                            "session": sess_name,
                            "taught": [f"P{p+1}: {cells[p]}" for p in taught_periods],
                            "all_periods": [f"P{p+1}: {cells[p]}" for p in range(5)]
                        })

    total_gap2 += t_gap2
    total_gap1 += t_gap1
    total_singletons += t_singletons
    total_sessions += t_sessions

print("\n=== EXCEL TIMETABLE OVERALL METRICS ===")
print(f"Total Sessions (tsBuoiDay): {total_sessions}")
print(f"Total 1-Period Sessions (soBuoiDay1): {total_singletons}")
print(f"Total 1-Period Gaps (soBuoiTrong1): {total_gap1}")
print(f"Total 2-Period Gaps (soBuoiTrong2): {total_gap2}")

print(f"\n=== DETAILS OF {len(gap2_details)} SESSIONS WITH 2-PERIOD GAPS ===")
for idx, g in enumerate(gap2_details, 1):
    print(f"\n{idx}. Teacher: {g['teacher']} ({g['day']} {g['session']})")
    print(f"   Taught: {g['taught']}")
    print(f"   Full Schedule: {g['all_periods']}")
