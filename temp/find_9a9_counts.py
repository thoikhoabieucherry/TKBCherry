import sys
import os
import openpyxl
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')

def find_unplaced_9a9():
    base_dir = r"C:\Users\Love\Documents\Codex"
    print("Checking all excel files in Codex for 9A9 / 9/9 lesson counts...")
    
    for root, dirs, files in os.walk(base_dir):
        if "node_modules" in root or ".git" in root or "TKBCherry-backups" in root:
            continue
        for f in files:
            if f.endswith(".xlsx") and not f.startswith("~$"):
                p = os.path.join(root, f)
                try:
                    wb = openpyxl.load_workbook(p, data_only=True)
                    for sname in wb.sheetnames:
                        ws = wb[sname]
                        # Count how many cells mention 9A9 or 9/9
                        count_9a9 = 0
                        count_9_9 = 0
                        for r in range(1, ws.max_row + 1):
                            for c in range(1, ws.max_column + 1):
                                v = str(ws.cell(r, c).value or "")
                                if "9A9" in v: count_9a9 += 1
                                if "9/9" in v: count_9_9 += 1
                        if count_9a9 > 0 or count_9_9 > 0:
                            print(f"File: {p} | Sheet: {sname} -> 9A9 count: {count_9a9}, 9/9 count: {count_9_9}")
                    wb.close()
                except:
                    pass

if __name__ == "__main__":
    find_unplaced_9a9()
