import sys
import openpyxl
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')

def parse_tonggv_file(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Sheet1']
    
    # 1. Parse teachers from Row 2
    teachers = {} # col_idx -> teacher_name
    for c in range(4, ws.max_column + 1):
        v = ws.cell(2, c).value
        if v is not None and str(v).strip() != "":
            teachers[c] = str(v).strip()
            
    # 2. Parse time slots from Rows 3 to ws.max_row
    # (day, session, period) -> row_idx
    slot_to_row = {}
    row_to_slot = {}
    
    current_day = ""
    current_sess = ""
    
    for r in range(3, ws.max_row + 1):
        day_val = ws.cell(r, 1).value
        sess_val = ws.cell(r, 2).value
        period_val = ws.cell(r, 3).value
        
        if day_val is not None and str(day_val).strip() != "":
            current_day = str(day_val).strip()
        if sess_val is not None and str(sess_val).strip() != "":
            current_sess = str(sess_val).strip()
            
        p_str = str(period_val).strip() if period_val is not None else ""
        if p_str != "":
            slot = (current_day, current_sess, p_str)
            slot_to_row[slot] = r
            row_to_slot[r] = slot
            
    # 3. Build teacher schedule matrix: (teacher, day, session, period) -> cell_str
    # and class schedule matrix: (class, day, session, period) -> (teacher, subject, raw_str)
    gv_grid = {}
    lop_grid = {} # (cls, day, sess, period) -> (teacher, subject)
    
    for r, (day, sess, p_str) in row_to_slot.items():
        for c, teacher in teachers.items():
            val = ws.cell(r, c).value
            raw_str = str(val).strip() if val is not None else ""
            gv_grid[(teacher, day, sess, p_str)] = raw_str
            if raw_str != "":
                # parse class and subject from raw_str, e.g. "6A13 - HĐTN 1" or "9A4 - Văn"
                parts = raw_str.split(" - ")
                cls_name = parts[0].strip()
                subj_name = parts[1].strip() if len(parts) > 1 else ""
                
                # Check for class collision
                cls_key = (cls_name, day, sess, p_str)
                if cls_key in lop_grid:
                    print(f"WARNING: Class Collision at {cls_key}: {lop_grid[cls_key]} vs {(teacher, subj_name)}")
                lop_grid[cls_key] = (teacher, subj_name, raw_str)
                
    return teachers, slot_to_row, row_to_slot, gv_grid, lop_grid

def analyze_all_gaps(teachers, gv_grid, row_to_slot):
    days = ["2", "3", "4", "5", "6", "7"]
    sessions = ["Sáng", "Chiều"]
    
    gaps_2 = []
    gaps_1 = []
    gaps_3_plus = []
    
    teacher_gap_details = defaultdict(list)
    
    for teacher in teachers.values():
        for day in days:
            for sess in sessions:
                teaching = []
                for p in range(1, 6):
                    p_str = str(p)
                    val = gv_grid.get((teacher, day, sess, p_str), "")
                    if val != "":
                        teaching.append((p, val))
                if len(teaching) >= 2:
                    p_indices = [t[0] for t in teaching]
                    min_p = min(p_indices)
                    max_p = max(p_indices)
                    
                    cur_gap = 0
                    cur_start = 0
                    for p in range(min_p, max_p + 1):
                        val = gv_grid.get((teacher, day, sess, str(p)), "")
                        if val == "":
                            if cur_gap == 0:
                                cur_start = p
                            cur_gap += 1
                        else:
                            if cur_gap > 0:
                                gap_obj = {
                                    'teacher': teacher,
                                    'day': day,
                                    'sess': sess,
                                    'gap_length': cur_gap,
                                    'gap_periods': list(range(cur_start, cur_start + cur_gap)),
                                    'teaching': teaching
                                }
                                if cur_gap == 1:
                                    gaps_1.append(gap_obj)
                                elif cur_gap == 2:
                                    gaps_2.append(gap_obj)
                                else:
                                    gaps_3_plus.append(gap_obj)
                                teacher_gap_details[teacher].append(gap_obj)
                                cur_gap = 0
                    if cur_gap > 0:
                        gap_obj = {
                            'teacher': teacher,
                            'day': day,
                            'sess': sess,
                            'gap_length': cur_gap,
                            'gap_periods': list(range(cur_start, cur_start + cur_gap)),
                            'teaching': teaching
                        }
                        if cur_gap == 1:
                            gaps_1.append(gap_obj)
                        elif cur_gap == 2:
                            gaps_2.append(gap_obj)
                        else:
                            gaps_3_plus.append(gap_obj)
                        teacher_gap_details[teacher].append(gap_obj)
                        
    return gaps_1, gaps_2, gaps_3_plus, teacher_gap_details

if __name__ == "__main__":
    filepath = r"C:\Users\Love\Documents\Codex\MD\tonggv0417082026.xlsx"
    teachers, slot_to_row, row_to_slot, gv_grid, lop_grid = parse_tonggv_file(filepath)
    
    print(f"Total teachers: {len(teachers)}")
    print(f"Total time slots: {len(slot_to_row)}")
    print(f"Total teaching assignments: {len(lop_grid)}")
    
    gaps_1, gaps_2, gaps_3_plus, teacher_gap_details = analyze_all_gaps(teachers, gv_grid, row_to_slot)
    
    print("\n" + "="*70)
    print("ANALYSIS OF GAPS IN tonggv0417082026.xlsx")
    print("="*70)
    print(f"Total 2-Period Gaps (Tiết trống 2 tiết): {len(gaps_2)}")
    print(f"Total 1-Period Gaps (Tiết trống 1 tiết): {len(gaps_1)}")
    print(f"Total >=3-Period Gaps (Tiết trống >=3 tiết): {len(gaps_3_plus)}")
    print("="*70)
    
    print("\n--- DETAILED 2-PERIOD GAPS ---")
    for i, g in enumerate(gaps_2, 1):
        print(f"{i:2d}. Teacher: {g['teacher']:<15} | Thứ {g['day']} ({g['sess']}) | Trống tiết: {g['gap_periods']} | Đang dạy: {g['teaching']}")
