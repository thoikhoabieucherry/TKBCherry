from __future__ import annotations

from pathlib import Path
from typing import Any
import re
import unicodedata

import openpyxl


def _workbook_roots(root: Path) -> tuple[Path, ...]:
    """Return compatible workbook locations for packaged and source layouts."""

    candidates = [
        root,
        root / "data",
        root.parent / "data",
        root.parent.parent / "data",
    ]
    out: list[Path] = []
    seen: set[str] = set()
    for candidate in candidates:
        key = str(candidate.resolve(strict=False)).casefold()
        if key in seen:
            continue
        seen.add(key)
        out.append(candidate)
    return tuple(out)


def _resolve_workbook(
    root: Path,
    *names: str,
    patterns: tuple[str, ...] = (),
) -> Path:
    searched: list[str] = []
    for directory in _workbook_roots(root):
        for name in names:
            candidate = directory / name
            searched.append(str(candidate))
            if candidate.is_file():
                return candidate
        for pattern in patterns:
            searched.append(str(directory / pattern))
            matches = sorted(path for path in directory.glob(pattern) if path.is_file())
            if matches:
                return matches[-1]
    label = ", ".join(names) or ", ".join(patterns)
    raise FileNotFoundError(f"Khong tim thay workbook {label}. Da tim: {'; '.join(searched)}")


def _rows(workbook_path: Path) -> list[tuple[Any, ...]]:
    workbook = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        return [tuple(row) for row in workbook.active.iter_rows(values_only=True)]
    finally:
        workbook.close()


def _norm_key(value: Any) -> str:
    text = str(value or "").strip()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    return re.sub(r"[^0-9a-z]+", "", text.casefold())


def _class_canonical(value: Any) -> str:
    text = str(value or "").strip()
    compact = re.sub(r"\s+", "", text)
    match = re.fullmatch(r"(\d+)[aA](\d+)", compact)
    if match:
        return f"{int(match.group(1))}/{int(match.group(2))}"
    match = re.fullmatch(r"(\d+)\s*/\s*(\d+)", text)
    if match:
        return f"{int(match.group(1))}/{int(match.group(2))}"
    return text


def _slot_from_excel_header(value: Any) -> str | None:
    text = str(value or "").strip()
    direct = re.fullmatch(r"thu(\d+)\|(sang|chieu)\|(\d+)", text, re.IGNORECASE)
    if direct:
        return f"thu{int(direct.group(1))}|{direct.group(2).lower()}|{int(direct.group(3))}"
    match = re.fullmatch(r"T(\d+)_(S|C)(\d+)", text, re.IGNORECASE)
    if not match:
        return None
    day = int(match.group(1))
    session = "sang" if match.group(2).upper() == "S" else "chieu"
    period = int(match.group(3)) - 1
    if day < 2 or day > 7 or period < 0:
        return None
    return f"thu{day}|{session}|{period}"


def _truthy_excel(value: Any) -> bool:
    if value is True:
        return True
    if isinstance(value, (int, float)):
        return value != 0
    key = _norm_key(value)
    return key in {"1", "x", "xx", "nghi", "off", "true", "yes", "y", "co"}


def _class_alias_map(classes: list[dict[str, Any]]) -> dict[str, str]:
    out: dict[str, str] = {}

    def add(alias: Any, class_id: str) -> None:
        for candidate in {str(alias or "").strip(), _class_canonical(alias)}:
            key = _norm_key(candidate)
            if key and key not in out:
                out[key] = class_id

    for index, cls in enumerate(classes, start=1):
        class_id = str(cls.get("id") or cls.get("ten") or "").strip()
        if not class_id:
            continue
        add(class_id, class_id)
        add(cls.get("ten"), class_id)
        add(cls.get("ten2"), class_id)
        add(f"L{index:03d}", class_id)
    return out


def _load_class_fixed_off_excel(
    root: Path,
    classes: list[dict[str, Any]],
) -> tuple[dict[str, dict[str, bool]], dict[str, list[str]], str | None]:
    try:
        path = _resolve_workbook(
            root,
            "tiet_nghi_lop.xlsx",
            "co_dinh_tiet_lop.xlsx",
            patterns=("co_dinh_tiet_lop_*.xlsx",),
        )
    except FileNotFoundError:
        return {}, {}, None

    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook["TietNghiLop"] if "TietNghiLop" in workbook.sheetnames else workbook.active
        rows = [tuple(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()

    alias_map = _class_alias_map(classes)
    header_index = -1
    slot_columns: list[tuple[int, str]] = []
    for row_index, row in enumerate(rows):
        found = [(idx, slot) for idx, value in enumerate(row) if (slot := _slot_from_excel_header(value))]
        if found:
            header_index = row_index
            slot_columns = found
            break
    if header_index < 0 or not slot_columns:
        return {}, {}, path.name

    header = rows[header_index]
    class_col = next(
        (
            idx
            for idx, value in enumerate(header)
            if _norm_key(value) in {"malop", "lop", "classid", "class", "id", "ma"}
        ),
        -1,
    )
    name_col = next(
        (
            idx
            for idx, value in enumerate(header)
            if _norm_key(value) in {"tenlop", "tenlophoc", "lopname", "classname", "name"}
        ),
        -1,
    )
    if class_col < 0 and name_col < 0:
        return {}, {}, path.name

    fixed_off: dict[str, dict[str, bool]] = {}
    user_off: dict[str, list[str]] = {}
    for row in rows[header_index + 1 :]:
        raw_values = []
        if class_col >= 0 and class_col < len(row):
            raw_values.append(row[class_col])
        if name_col >= 0 and name_col < len(row):
            raw_values.append(row[name_col])
        class_id = ""
        for raw in raw_values:
            class_id = alias_map.get(_norm_key(raw)) or alias_map.get(_norm_key(_class_canonical(raw))) or ""
            if class_id:
                break
        if not class_id:
            continue
        slots = [
            slot
            for idx, slot in slot_columns
            if idx < len(row) and _truthy_excel(row[idx])
        ]
        if slots:
            fixed_off[class_id] = {slot: True for slot in slots}
            user_off[class_id] = slots
    return fixed_off, user_off, path.name


def _header_column(
    header: tuple[Any, ...],
    aliases: set[str],
    fallback: int,
) -> int:
    for index, value in enumerate(header):
        if _norm_key(value) in aliases:
            return index
    return fallback


def _row_value(row: tuple[Any, ...], index: int) -> Any:
    return row[index] if 0 <= index < len(row) else None


def build_ui_fixture_from_workbooks(web_root: str | Path, *, include_fixed_off_excel: bool = False) -> dict[str, Any]:
    root = Path(web_root)
    lop_rows = _rows(_resolve_workbook(root, "lop.xlsx"))
    gv_rows = _rows(_resolve_workbook(root, "gv.xlsx", "giaovien.xlsx"))
    mon_rows = _rows(_resolve_workbook(root, "mon.xlsx", "monhoc.xlsx"))
    tiet_rows = _rows(_resolve_workbook(root, "tietchuan.xlsx"))
    pccm_rows = _rows(_resolve_workbook(root, "pccm.xlsx", "PCCM.xlsx"))

    lop_header = lop_rows[0] if lop_rows else ()
    class_col = _header_column(
        lop_header,
        {"malop", "tenlop", "lop", "classid", "classname"},
        1,
    )
    grade_col = _header_column(lop_header, {"khoi", "khoihoc", "grade"}, 3)
    alias_col = _header_column(
        lop_header,
        {"tenlop2", "tentat", "tenkhac", "alias", "classalias"},
        -1,
    )
    session_col = _header_column(lop_header, {"buoi", "buoihoc", "session"}, 4)
    location_col = _header_column(lop_header, {"diadiem", "location", "phong"}, 5)

    classes = [
        {
            "id": str(_row_value(row, class_col)).strip(),
            "ten": str(_row_value(row, class_col)).strip(),
            "ten2": str(_row_value(row, alias_col)).strip() if _row_value(row, alias_col) else "",
            "khoi": str(_row_value(row, grade_col)).strip(),
            "buoi": str(_row_value(row, session_col)).strip() if _row_value(row, session_col) else "",
            "diadiem": str(_row_value(row, location_col)).strip() if _row_value(row, location_col) else "",
        }
        for row in lop_rows[1:]
        if _row_value(row, class_col) and _row_value(row, grade_col)
    ]
    teachers = [
        {
            "hodem": str(row[1]).strip() if row[1] else "",
            "ten": str(row[2]).strip() if row[2] else "",
            "magv": str(row[3]).strip() if row[3] else "",
        }
        for row in gv_rows[1:]
        if len(row) >= 4 and row[3]
    ]
    subjects = [
        {
            "ten": str(row[1]).strip() if row[1] else "",
            "ma": str(row[2]).strip() if len(row) > 2 and row[2] else "",
            "ma2": str(row[3]).strip() if len(row) > 3 and row[3] else "",
        }
        for row in mon_rows[1:]
        if len(row) >= 2 and row[1]
    ]
    periods = [
        {
            "khoi": str(row[1]).strip(),
            "ten": str(row[2]).strip(),
            "sotiet": row[3],
            "gioihan": row[4],
        }
        for row in tiet_rows[1:]
        if len(row) >= 5 and row[1] and row[2]
    ]

    pccm_matrix: dict[str, str] = {}
    headers = [str(value).strip() if value else "" for value in pccm_rows[0][1:]]
    for row in pccm_rows[1:]:
        if not row or not row[0]:
            continue
        class_name = str(row[0]).strip()
        for subject, teacher in zip(headers, row[1:]):
            if subject and teacher:
                pccm_matrix[f"{class_name}|{subject}"] = str(teacher).strip()

    data: dict[str, Any] = {
        "lop": classes,
        "giaovien": teachers,
        "monhoc": subjects,
        "mon": periods,
        "pccmMatrix": pccm_matrix,
        "tkbConstraints": {},
    }
    if include_fixed_off_excel:
        class_fixed_off, user_off, fixed_off_workbook = _load_class_fixed_off_excel(root, classes)
        if class_fixed_off:
            data["tkbUserOff"] = user_off
            data["tkbConstraints"] = {
                "fixedOff": {
                    "class": class_fixed_off,
                    "teacher": {},
                    "subject": {},
                    "room": {},
                    "subjectGroup": {},
                }
            }
            data["sampleMeta"] = {
                "fixedOffExcel": fixed_off_workbook,
                "fixedOffClassCount": len(class_fixed_off),
                "fixedOffSlotCount": sum(len(slots) for slots in class_fixed_off.values()),
            }
    return data
