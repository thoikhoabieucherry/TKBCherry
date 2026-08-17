import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook("C:/Users/Love/Documents/Codex/phancong/tesst/da xep.xlsx", data_only=True)
sheet = wb.active

# Let's inspect header rows (rows 1-5) to see class names and structure
print("Header inspection:")
for r in range(1, 6):
    row_vals = [sheet.cell(r, c).value for c in range(1, 35)]
    print(f"Row {r}: {row_vals}")

# Let's see how many classes there are across columns
classes = []
for c in range(3, sheet.max_column + 1, 2):
    c1 = sheet.cell(4, c).value
    c2 = sheet.cell(4, c+1).value
    if c1 or c2:
        classes.append((c, c1 or c2))

print(f"\nFound {len(classes)} classes:", classes)

# Let's check a sample class schedule (e.g. column 3/4)
# Rows 6-35 represent days and periods
print("\nSample Schedule for Class 1 (cols 3 & 4):")
for r in range(6, sheet.max_row + 1):
    day = sheet.cell(r, 1).value
    period = sheet.cell(r, 2).value
    sang = sheet.cell(r, 3).value
    chieu = sheet.cell(r, 4).value
    if any([day, period, sang, chieu]):
        print(f"Thứ {day or ' '} Tiết {period or ' '}: Sáng = [{sang or ''}] | Chiều = [{chieu or ''}]")
