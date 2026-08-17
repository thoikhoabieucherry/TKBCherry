import openpyxl
import os
import sys

# Ensure UTF-8 output
sys.stdout.reconfigure(encoding='utf-8')

base_dir = r"C:\Users\Love\Documents\Codex\temp"
files = ["xepmoi.xlsx", "1.xlsx", "2.xlsx", "3.xlsx"]

def inspect_file(wb_path):
    wb = openpyxl.load_workbook(wb_path, data_only=True)
    print(f"\n=======================================================")
    print(f"File: {os.path.basename(wb_path)}")
    print(f"Sheets: {wb.sheetnames}")
    for name in wb.sheetnames:
        sh = wb[name]
        print(f"  Sheet '{name}': {sh.max_row} rows x {sh.max_column} cols")
        # Print header rows
        for r in range(1, min(7, sh.max_row + 1)):
            row_vals = [str(sh.cell(r, c).value or "") for c in range(1, min(20, sh.max_column + 1))]
            print(f"    R{r}: {row_vals}")

for f in files:
    inspect_file(os.path.join(base_dir, f))
