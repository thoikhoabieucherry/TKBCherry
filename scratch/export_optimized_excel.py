import openpyxl, copy, sys
sys.stdout.reconfigure(encoding='utf-8')
import analyze_temp_files as atf
from compare_2_approaches import h1_t_grid, h1_c_grid, base_m, h1_m_final

# Load original workbook to preserve styles, widths, and layout
wb = openpyxl.load_workbook(r'C:\Users\Love\Documents\Codex\TKBCherry\temp\tonggv0318082026.xlsx')
ws = wb['Sheet1']

# Get teacher columns
teachers = []
for c in range(4, ws.max_column + 1):
    val = ws.cell(2, c).value
    if val:
        teachers.append((c, str(val).strip()))

# Write optimized schedule to sheet
for r in range(3, 63):
    slot_idx = r - 3
    for col_idx, tname in teachers:
        opt_val = h1_t_grid.get(tname, [None]*60)[slot_idx]
        ws.cell(r, col_idx).value = opt_val if opt_val is not None else ""

output_path = r'C:\Users\Love\Documents\Codex\TKBCherry\temp\tonggv0318082026_optimized.xlsx'
wb.save(output_path)
print(f"Exported optimized timetable to: {output_path}")

print(f"\nAudit results:")
print(f"- Singletons: {base_m['singletons']} -> {h1_m_final['singletons']}")
print(f"- Total teacher sessions: {base_m['total_sessions']} -> {h1_m_final['total_sessions']}")
print(f"- Off-period violations: {h1_m_final['off_violations']}")
