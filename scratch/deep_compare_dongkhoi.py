import openpyxl
import os
import sys
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')

base_dir = r"C:\Users\Love\Documents\Codex\temp"
files = ["xepmoi.xlsx", "1.xlsx", "2.xlsx", "3.xlsx"]

DAYS = ["2", "3", "4", "5", "6", "7"]
PERIODS = 5

def parse_schedule(wb_path):
    wb = openpyxl.load_workbook(wb_path, data_only=True)
    sh = wb["TKB_LOP_SC"]
    
    # Col 3 onwards: Classes
    # Row 4: Class names, merged across 2 columns (Sang, Chieu)
    # Row 5: Sang / Chieu
    # Row 6..35: Days & Periods
    classes = {}
    
    current_class = ""
    col_to_class_session = {} # col_idx -> (class_name, session: 0 for Sang, 1 for Chieu)
    
    for c in range(3, sh.max_column + 1):
        c_name = sh.cell(4, c).value
        if c_name:
            current_class = str(c_name).strip()
        sess_str = str(sh.cell(5, c).value or "").strip().lower()
        sess = 0 if "sáng" in sess_str or "sang" in sess_str else 1
        if current_class:
            col_to_class_session[c] = (current_class, sess)
            if current_class not in classes:
                classes[current_class] = {}

    # Read rows 6 to 35
    # (day, period)
    # Row 6: Day 2, P1
    # Row 10: Day 2, P5
    # Row 11: Day 3, P1 ...
    teacher_schedule = defaultdict(lambda: defaultdict(list)) # teacher -> (day, session) -> list of (period, mon, class)
    class_schedule = defaultdict(lambda: defaultdict(list))   # class -> (day, session) -> list of (period, mon, teacher)
    all_cells = []

    for r in range(6, 36):
        d_val = str(sh.cell(r, 1).value or "").strip()
        p_val = str(sh.cell(r, 2).value or "").strip()
        
        # Calculate day index and period index
        r_idx = r - 6
        d_idx = r_idx // 5 # 0..5 (Day 2..7)
        p_idx = r_idx % 5  # 0..4 (P1..P5)
        day_str = DAYS[d_idx]
        
        for c, (c_name, sess) in col_to_class_session.items():
            cell_val = str(sh.cell(r, c).value or "").strip()
            if cell_val and cell_val != "-" and cell_val != "OFF":
                # cell_val format: "Mon - Teacher" e.g. "ChCờ - P.My" or "KHTN - G.Lâm"
                parts = cell_val.split("-")
                mon = parts[0].strip()
                gv = parts[1].strip() if len(parts) > 1 else ""
                
                teacher_schedule[gv][(day_str, sess)].append((p_idx, mon, c_name))
                class_schedule[c_name][(day_str, sess)].append((p_idx, mon, gv))
                all_cells.append({
                    "class": c_name, "session": sess, "day": day_str, "p": p_idx,
                    "mon": mon, "gv": gv, "raw": cell_val
                })

    # Calculate metrics for teachers
    singletons = 0
    total_sessions = 0
    total_days = 0
    gaps_1 = 0
    gaps_2 = 0
    gaps_3plus = 0
    
    teacher_stats = {}

    for gv, sess_map in teacher_schedule.items():
        if not gv or gv == "None": continue
        gv_days = set()
        gv_singletons = []
        gv_sessions = 0
        
        for (day, sess), items in sess_map.items():
            k = len(items)
            if k > 0:
                gv_sessions += 1
                gv_days.add(day)
                if k == 1:
                    singletons += 1
                    gv_singletons.append((day, sess, items[0]))
                elif k > 1:
                    items.sort(key=lambda x: x[0])
                    p_min = items[0][0]
                    p_max = items[-1][0]
                    span = p_max - p_min + 1
                    gap = span - k
                    if gap == 1: gaps_1 += 1
                    elif gap == 2: gaps_2 += 1
                    elif gap >= 3: gaps_3plus += 1
                    
        total_sessions += gv_sessions
        total_days += len(gv_days)
        teacher_stats[gv] = {
            "sessions": gv_sessions,
            "days": len(gv_days),
            "singletons": gv_singletons
        }

    return {
        "file": os.path.basename(wb_path),
        "classes_count": len(classes),
        "teachers_count": len(teacher_schedule),
        "cells_count": len(all_cells),
        "singletons": singletons,
        "total_sessions": total_sessions,
        "total_days": total_days,
        "gaps_1": gaps_1,
        "gaps_2": gaps_2,
        "gaps_3plus": gaps_3plus,
        "teacher_stats": teacher_stats,
        "teacher_schedule": teacher_schedule,
        "class_schedule": class_schedule,
        "all_cells": all_cells
    }

results = [parse_schedule(os.path.join(base_dir, f)) for f in files]

print(f"{'File':<12} | {'Singletons (Dạy 1)':<18} | {'Sessions (Buổi dạy)':<20} | {'Days (Ngày dạy)':<16} | {'Gaps 1':<8} | {'Gaps 2':<8}")
print("-" * 90)
for r in results:
    print(f"{r['file']:<12} | {r['singletons']:<18} | {r['total_sessions']:<20} | {r['total_days']:<16} | {r['gaps_1']:<8} | {r['gaps_2']:<8}")
