import openpyxl
import os
import sys

# Force UTF-8 output on Windows
sys.stdout.reconfigure(encoding='utf-8')

path_da_xep = "C:/Users/Love/Documents/Codex/phancong/tesst/da xep.xlsx"
path_nghi = "C:/Users/Love/Documents/Codex/phancong/tesst/nghi.xlsx"

def inspect_excel(path):
    print(f"\n=================== {os.path.basename(path)} ===================")
    wb = openpyxl.load_workbook(path, data_only=True)
    print("Sheets:", wb.sheetnames)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\n--- Sheet: {sheet_name} (max_row={sheet.max_row}, max_col={sheet.max_column}) ---")
        for r in range(1, min(35, sheet.max_row + 1)):
            row_vals = [str(sheet.cell(r, c).value or "").strip() for c in range(1, min(25, sheet.max_column + 1))]
            if any(row_vals):
                print(f"Row {r:2d}: {row_vals}")

inspect_excel(path_nghi)
inspect_excel(path_da_xep)
