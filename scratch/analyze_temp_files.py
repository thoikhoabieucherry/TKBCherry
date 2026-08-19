import openpyxl, sys, json
sys.stdout.reconfigure(encoding='utf-8')

# 1. Parse co_dinh_tiet_lop_20260818.xlsx
wb_off = openpyxl.load_workbook(r'C:\Users\Love\Documents\Codex\TKBCherry\temp\co_dinh_tiet_lop_20260818.xlsx', data_only=True)
ws_off = wb_off['CoDinhLop']

classes = []
for c in range(4, ws_off.max_column + 1):
    val = ws_off.cell(1, c).value
    if val:
        classes.append((c, str(val).strip()))

print(f'Found {len(classes)} classes in co_dinh_tiet_lop')

class_off_slots = {cname: set() for _, cname in classes}
slot_info = []

curr_day = None
curr_session = None

for r in range(2, 62):
    day_val = ws_off.cell(r, 1).value
    session_val = ws_off.cell(r, 2).value
    period_val = ws_off.cell(r, 3).value
    
    if day_val:
        curr_day = str(day_val).replace('THỨ', '').strip()
    if session_val:
        curr_session = str(session_val).strip()
    
    period = int(period_val) if period_val is not None else 1
    slot_idx = r - 2 # 0..59
    slot_info.append({
        'slot': slot_idx,
        'day': curr_day,
        'session': curr_session,
        'period': period,
        'row': r
    })
    
    for col_idx, cname in classes:
        val = ws_off.cell(r, col_idx).value
        if val is not None and str(val).strip().lower() in ['nghỉ', 'nghi', 'x', '1', 'off']:
            class_off_slots[cname].add(slot_idx)

# 2. Parse tonggv0318082026.xlsx
wb_gv = openpyxl.load_workbook(r'C:\Users\Love\Documents\Codex\TKBCherry\temp\tonggv0318082026.xlsx', data_only=True)
ws_gv = wb_gv['Sheet1']

teachers = []
for c in range(4, ws_gv.max_column + 1):
    val = ws_gv.cell(2, c).value
    if val:
        teachers.append((c, str(val).strip()))

print(f'Found {len(teachers)} teachers in tonggv')

teacher_grid = {tname: [None]*60 for _, tname in teachers}
class_grid = {}

conflicts = []

for r in range(3, 63):
    slot_idx = r - 3
    for col_idx, tname in teachers:
        val = ws_gv.cell(r, col_idx).value
        if val is not None and str(val).strip():
            cell_text = str(val).strip()
            teacher_grid[tname][slot_idx] = cell_text
            
            parts = cell_text.split('-')
            cname = parts[0].strip()
            subj = parts[1].strip() if len(parts) > 1 else ''
            
            if cname not in class_grid:
                class_grid[cname] = [None]*60
            if class_grid[cname][slot_idx] is not None:
                conflicts.append((cname, slot_idx, class_grid[cname][slot_idx], (tname, subj, cell_text)))
            else:
                class_grid[cname][slot_idx] = (tname, subj, cell_text)

print(f'Total classes in timetable: {len(class_grid)}')
print(f'Total teacher-class overlap conflicts: {len(conflicts)}')
for conf in conflicts:
    print(f'  Overlap conflict: Class {conf[0]} Slot {conf[1]} -> {conf[2]} vs {conf[3]}')

# Check violations against class_off_slots
violations = []
for cname, sched in class_grid.items():
    off_set = class_off_slots.get(cname, set())
    for s_idx in range(60):
        if sched[s_idx] is not None and s_idx in off_set:
            info = slot_info[s_idx]
            tname, subj, raw = sched[s_idx]
            violations.append({
                'class': cname,
                'slot': s_idx,
                'day': info['day'],
                'session': info['session'],
                'period': info['period'],
                'teacher': tname,
                'subject': subj,
                'text': raw
            })

print(f'\nTotal class off-period violations in current schedule: {len(violations)}')
for v in violations:
    print(f"  Violation: Class {v['class']} on Thứ {v['day']} {v['session']} Tiết {v['period']} (Slot {v['slot']}) -> Teacher {v['teacher']}, Subj {v['subject']}")

# Teacher singletons (1 tiet / buoi)
print('\n=== TEACHER SINGLETONS (1 tiết/buổi) ===')
singletons = []
gaps = []
teacher_stats = {}

for tname, sched in teacher_grid.items():
    total_periods = sum(1 for x in sched if x is not None)
    sessions_taught = 0
    t_singletons = []
    t_gaps = []
    
    for d in range(6):
        for b in range(2):
            s_start = d * 10 + b * 5
            session_slots = [sched[s_start + p] for p in range(5)]
            taught = [(p, s_start + p, session_slots[p]) for p in range(5) if session_slots[p] is not None]
            if len(taught) > 0:
                sessions_taught += 1
            if len(taught) == 1:
                p, s_idx, text = taught[0]
                info = slot_info[s_idx]
                item = {
                    'teacher': tname,
                    'day': info['day'],
                    'session': info['session'],
                    'period': info['period'],
                    'slot': s_idx,
                    'text': text
                }
                singletons.append(item)
                t_singletons.append(item)
            elif len(taught) > 1:
                periods = [x[0] for x in taught]
                if max(periods) - min(periods) + 1 > len(taught):
                    item = {
                        'teacher': tname,
                        'day': slot_info[s_start]['day'],
                        'session': slot_info[s_start]['session'],
                        'periods': [p+1 for p in periods],
                        'taught': taught
                    }
                    gaps.append(item)
                    t_gaps.append(item)
                    
    teacher_stats[tname] = {
        'total_periods': total_periods,
        'sessions_taught': sessions_taught,
        'singletons': len(t_singletons),
        'gaps': len(t_gaps)
    }

print(f'Total Teacher Singletons (1 tiết/buổi): {len(singletons)}')
for s in singletons:
    print(f"  Singleton: GV {s['teacher']} -> Thứ {s['day']} {s['session']} Tiết {s['period']} ({s['text']})")

print(f'\nTotal Teacher Gaps (Tiết trống xen kẽ): {len(gaps)}')
for g in gaps:
    print(f"  Gap: GV {g['teacher']} -> Thứ {g['day']} {g['session']}, Tiết {g['periods']}")

# Check total periods taught by teacher vs singletons
print('\nTeachers with singletons:')
for tname, stat in teacher_stats.items():
    if stat['singletons'] > 0 or stat['gaps'] > 0:
        print(f"  GV {tname}: {stat['total_periods']} tiết, {stat['sessions_taught']} buổi, {stat['singletons']} tiết lẻ, {stat['gaps']} buổi lủng")
