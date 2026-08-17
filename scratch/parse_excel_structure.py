import sys
import openpyxl
import json

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r"C:\Users\Love\Documents\Codex\TKBCherry\temp\tonggv0917082026.xlsx", data_only=True)
ws = wb.active

print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")

# Row 2 contains teacher names
teachers = []
for c in range(4, ws.max_column + 1):
    val = ws.cell(2, c).value
    if val:
        teachers.append((c, str(val).strip()))

print(f"Total teachers found: {len(teachers)}")
print("Sample teachers:", teachers[:10])

# Days and sessions
# Row 3 to 62: 6 days * 10 periods = 60 rows
days = ["thu2", "thu3", "thu4", "thu5", "thu6", "thu7"]
# Check rows for day, session, period
row_mapping = []
current_day = 0
current_buoi = "sang"
for r in range(3, ws.max_row + 1):
    day_cell = ws.cell(r, 1).value
    buoi_cell = ws.cell(r, 2).value
    tiet_cell = ws.cell(r, 3).value
    if day_cell:
        # e.g. '2', '3', '4', '5', '6', '7'
        d_num = int(str(day_cell).strip())
        current_day = d_num - 2
    if buoi_cell:
        b_str = str(buoi_cell).strip().lower()
        current_buoi = "sang" if "sáng" in b_str or "sang" in b_str else "chieu"
    if tiet_cell is not None:
        try:
            t_num = int(str(tiet_cell).strip()) - 1
            row_mapping.append((r, days[current_day], current_buoi, t_num))
        except:
            pass

print(f"Total time slots mapped: {len(row_mapping)}")

# Build class TKB and teachers list
tkb = {}
classes_set = set()
pccm = {}

for r, thu, buoi, ti in row_mapping:
    for col, gv in teachers:
        val = ws.cell(r, col).value
        if val and str(val).strip():
            cell_str = str(val).strip()
            # Format: '6/10 - KHTN' or '6/7 - NDGDCĐP'
            if " - " in cell_str:
                cls, mon = cell_str.split(" - ", 1)
                cls = cls.strip()
                mon = mon.strip()
            else:
                cls = cell_str
                mon = "Mon"
            
            classes_set.add(cls)
            if cls not in tkb:
                tkb[cls] = {d: {"sang": ["" for _ in range(5)], "chieu": ["" for _ in range(5)]} for d in days}
            
            tkb[cls][thu][buoi][ti] = f"{mon} - {gv}"
            pccm[f"{cls}|{mon}"] = gv

lop_list = [{"id": c, "name": c, "buoi": "sang"} for c in sorted(classes_set)]
gv_list = [{"magv": g, "ten": g} for _, g in teachers]

data = {
    "lop": lop_list,
    "giaovien": gv_list,
    "tkb": tkb,
    "pccmMatrix": pccm
}

with open(r"C:\Users\Love\Documents\Codex\TKBCherry\scratch\excel_data_parsed.json", "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print("Saved scratch/excel_data_parsed.json successfully!")
print(f"Total classes: {len(classes_set)}")
