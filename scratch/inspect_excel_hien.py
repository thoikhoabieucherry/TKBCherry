import sys
import openpyxl
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r"C:\Users\Love\Documents\Codex\TKBCherry\temp\tonggv0917082026.xlsx", data_only=True)
print("Sheet names:", wb.sheetnames)

ws = wb.active
for r in range(1, 100):
    row_vals = [ws.cell(r, c).value for c in range(1, 20)]
    if any("Hiền" in str(v) or "6/10" in str(v) or "6A10" in str(v) for v in row_vals if v):
        print(f"Row {r}: {row_vals}")

