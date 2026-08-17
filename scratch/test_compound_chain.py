import sys
import openpyxl
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

file_path = Path(r"C:\Users\Love\Documents\Codex\TKBCherry\temp\tonggv0917082026.xlsx")
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb.active

# Map all teachers and classes into matrices
teachers = {}
for c in range(4, sheet.max_column + 1):
    gv = sheet.cell(2, c).value
    if gv:
        teachers[str(gv).strip()] = c

# Build 60-slot grid for all teachers
teacher_grids = {}
for gv, col in teachers.items():
    grid = []
    for r in range(3, 63):
        v = sheet.cell(r, col).value
        grid.append(str(v).strip() if v is not None else "")
    teacher_grids[gv] = grid

# Build 60-slot grid for all classes
class_grids = {}
# Find all classes
all_classes = set()
for gv, grid in teacher_grids.items():
    for slot_val in grid:
        if " - " in slot_val:
            cls = slot_val.split(" - ")[0].strip()
            all_classes.add(cls)

for cls in all_classes:
    cgrid = [""] * 60
    for gv, grid in teacher_grids.items():
        for s, slot_val in enumerate(grid):
            if slot_val.startswith(f"{cls} - "):
                cgrid[s] = f"{slot_val} ({gv})"
    class_grids[cls] = cgrid

print(f"Total teachers: {len(teacher_grids)}, Total classes: {len(class_grids)}")

# Find compound 2-hop moves for TD.Kiệt (7A18 - GDTC at slot 15: Tue PM T1)
print("\n" + "="*70)
print("SEARCHING VALID 2-HOP COMPOUND MOVES FOR TD.Kiệt (7A18 - GDTC, slot 15):")
print("="*70)

# TD.Kiệt wants to move act1 (7A18 GDTC) from slot 15 to slot s2 in Fri PM (slots 45..49) or Wed PM (slots 25..29)
gdtc_target_slots = [25, 26, 27, 28, 29, 45, 46, 47, 48, 49]
cgrid_7A18 = class_grids["7A18"]
kiet_grid = teacher_grids["TD.Kiệt"]

found_moves_kiet = []
for s2 in gdtc_target_slots:
    if kiet_grid[s2] != "": continue # Kiệt must be free at s2
    
    # 7A18 at s2 has act2
    act2_str = cgrid_7A18[s2]
    if not act2_str or "(" not in act2_str: continue
    
    subj2 = act2_str.split(" (")[0]
    gv2 = act2_str.split(" (")[1].rstrip(")")
    gv2_grid = teacher_grids.get(gv2)
    if not gv2_grid: continue
    
    # Can act2 move to ANY other slot s3 where 7A18 is empty and gv2 is free?
    for s3 in range(60):
        if s3 == 15 or s3 == s2: continue
        if cgrid_7A18[s3] == "" and gv2_grid[s3] == "":
            found_moves_kiet.append((s2, act2_str, gv2, s3))
            print(f" -> FOUND: Move {subj2} (GV: {gv2}) from slot {s2} to slot {s3}; then Move 7A18-GDTC (TD.Kiệt) from slot 15 to slot {s2}!")

# Find compound 2-hop moves for T.Huy (7A17 - Toán at slot 55: Sat PM T1)
print("\n" + "="*70)
print("SEARCHING VALID 2-HOP COMPOUND MOVES FOR T.Huy (7A17 - Toán, slot 55):")
print("="*70)

huy_target_slots = [5, 6, 7, 8, 9, 35, 36, 37, 38, 39]
cgrid_7A17 = class_grids["7A17"]
huy_grid = teacher_grids["T.Huy"]

found_moves_huy = []
for s2 in huy_target_slots:
    if huy_grid[s2] != "": continue # Huy must be free at s2
    
    act2_str = cgrid_7A17[s2]
    if not act2_str or "(" not in act2_str: continue
    if "HĐTN" in act2_str: continue # skip fixed
    
    subj2 = act2_str.split(" (")[0]
    gv2 = act2_str.split(" (")[1].rstrip(")")
    gv2_grid = teacher_grids.get(gv2)
    if not gv2_grid: continue
    
    for s3 in range(60):
        if s3 == 55 or s3 == s2: continue
        if cgrid_7A17[s3] == "" and gv2_grid[s3] == "":
            found_moves_huy.append((s2, act2_str, gv2, s3))
            print(f" -> FOUND: Move {subj2} (GV: {gv2}) from slot {s2} to slot {s3}; then Move 7A17-Toán (T.Huy) from slot 55 to slot {s2}!")

print(f"\nTotal valid 2-hop solutions: Kiệt={len(found_moves_kiet)}, Huy={len(found_moves_huy)}")

