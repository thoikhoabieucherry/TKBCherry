import sys
import json
import openpyxl
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

excel_path = Path(r"C:\Users\Love\Documents\Codex\TKBCherry\temp\tonggv0917082026.xlsx")
base_json_path = Path(r"C:\Users\Love\Documents\Codex\TKBCherry\scratch\default_school_0317.json")

with open(base_json_path, "r", encoding="utf-8") as f:
    data = json.load(f)

wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb.active

teachers = {}
for c in range(4, sheet.max_column + 1):
    gv = sheet.cell(2, c).value
    if gv: teachers[str(gv).strip()] = c

DAYS_LIST = ["thu2", "thu3", "thu4", "thu5", "thu6", "thu7"]
SESSIONS_LIST = ["sang", "chieu"]

tkb = {}
for lop_obj in data["lop"]:
    cId = lop_obj["id"]
    tkb[cId] = {}
    for d in DAYS_LIST:
        tkb[cId][d] = {
            "sang": [""] * 5,
            "chieu": [""] * 5
        }

for gv, col in teachers.items():
    for r in range(3, 63):
        slot_idx = r - 3
        d_idx = slot_idx // 10
        b_idx = (slot_idx % 10) // 5
        p_idx = slot_idx % 5
        
        thu = DAYS_LIST[d_idx]
        buoi = SESSIONS_LIST[b_idx]
        
        v = sheet.cell(r, col).value
        if v and " - " in str(v):
            parts = str(v).split(" - ")
            cId = parts[0].strip()
            subj = parts[1].strip()
            if cId in tkb:
                is_fixed = False
                if "HĐTN" in subj and (" 1" in subj or " 2" in subj):
                    is_fixed = True
                
                tkb[cId][thu][buoi][p_idx] = {
                    "mon": subj,
                    "gv": gv,
                    "val": f"{cId} - {subj}",
                    "cd": 1 if is_fixed else 0,
                    "isFixed": is_fixed
                }

data["tkb"] = tkb

out_path = Path(r"C:\Users\Love\Documents\Codex\TKBCherry\scratch\test_state_0917.json")
with open(out_path, "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print(f"Exported correct lowercase-key fixture to {out_path} ({len(tkb)} classes)")

