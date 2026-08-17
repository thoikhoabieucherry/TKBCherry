import sys
import openpyxl
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

file_path = Path(r"C:\Users\Love\Documents\Codex\TKBCherry\temp\tonggv0917082026.xlsx")
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb.active

# Let's inspect T.Huy (Col 26) and 7A17, 7A18
# T.Huy:
# Mon PM (slots 5..9):
#  s5 (T1): 7A17 - Toán
#  s6 (T2): 7A18 - Toán
#  s7 (T3): 7A18 - Toán
#  s8 (T4): 7A17 - HĐTN 2 (fixed)
#  s9 (T5): 7A17 - HĐTN 1 (fixed)
# Sat PM (slots 55..59):
#  s55 (T1): 7A17 - Toán (singleton!)

# Tue PM (slots 15..19): T.Huy is FREE.
# 7A18 on Tue PM (slots 15..19):
#  s15 (T1): GDTC (TD.Kiệt)
#  s16 (T2): MT (MT.Trang)
#  s17 (T3): GDĐP (CN.Chí)
#  s18 (T4): CN (CN.Liên)
#  s19 (T5): Nhạc (Nhạc.Dương)

# Notice: On Tue PM, 7A18 has 5 single periods of minor subjects!
# If 7A18 swaps [s16, s17] (MT, GDĐP) with [s6, s7] (Toán 7A18, Toán 7A18):
# Then on Tue PM: 7A18 has Toán (T2, T3) with T.Huy -> 2 periods (valid session for T.Huy, NO singleton!).
# And on Mon PM: T.Huy is FREE at s6, s7!
# Then 7A17 on Mon PM:
#  Currently at Mon PM: s5 (Toán), s6 (Văn - V.Lâm), s7 (Văn - V.Lâm), s8 (HĐTN), s9 (HĐTN).
#  Wait, 7A17 has Văn at s6, s7!
#  If 7A17 swaps s6 (Văn) with s55 (Toán from Sat PM):
#  Then 7A17 has Toán at s5, s6 (2 Math on Mon PM -> PERFECT!).
#  And Văn moves to s55 (Sat PM T1)!
#  Does V.Lâm teach another class at Sat PM T1?
#  Let's check V.Lâm schedule on Saturday!

# Let's check V.Lâm column (Col 22)
v_lam_col = 22
for r in range(3, 63):
    v = sheet.cell(r, v_lam_col).value
    if v:
        d_val = sheet.cell(r, 1).value
        b_val = sheet.cell(r, 2).value
        p_val = sheet.cell(r, 3).value
        print(f"V.Lâm: Thứ {d_val} {b_val} Tiết {p_val}: {v}")

