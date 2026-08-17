import sys
import openpyxl
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

file_path = Path(r"C:\Users\Love\Documents\Codex\TKBCherry\temp\tonggv0917082026.xlsx")
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb.active

teachers = {}
for c in range(4, sheet.max_column + 1):
    gv = sheet.cell(2, c).value
    if gv: teachers[str(gv).strip()] = c

teacher_grids = {}
for gv, col in teachers.items():
    grid = []
    for r in range(3, 63):
        v = sheet.cell(r, col).value
        grid.append(str(v).strip() if v is not None else "")
    teacher_grids[gv] = grid

# 7A17 schedule
cgrid_7A17 = [""] * 60
for gv, grid in teacher_grids.items():
    for s, val in enumerate(grid):
        if val.startswith("7A17 - "):
            cgrid_7A17[s] = f"{val} ({gv})"

print("7A17 on Tue PM (15..19):", cgrid_7A17[15:20])
print("7A17 on Wed PM (25..29):", cgrid_7A17[25:30])
print("7A17 on Fri PM (45..49):", cgrid_7A17[45:50])

# We want to place 7A17 Math at [t1, t2] in one of (Tue PM, Wed PM, Fri PM)
# And displace the two lessons at [t1, t2] to:
#  - slot 5 (Mon PM T1 - freed by 7A17 Math)
#  - slot 55 (Sat PM T1 - freed by 7A17 Math)
#  - or slot 59 (Sat PM T5 - currently empty in 7A17)

target_sessions = [
    ("Tue PM", list(range(15, 20))),
    ("Wed PM", list(range(25, 30))),
    ("Fri PM", list(range(45, 50))),
]

dest_slots = [5, 55, 59]

print("\n--- CHECKING ALL 2-FOR-2 / 2-FOR-3 COMBINATIONS FOR 7A17 TOÁN ---")
for sess_name, sess_slots in target_sessions:
    for i in range(len(sess_slots)):
        for j in range(i + 1, len(sess_slots)):
            s_a = sess_slots[i]
            s_b = sess_slots[j]
            
            act_a = cgrid_7A17[s_a]
            act_b = cgrid_7A17[s_b]
            
            if "HĐTN" in act_a or "HĐTN" in act_b: continue
            if "KHTN" in act_a and "KHTN" in act_b: continue # don't split double block
            
            gv_a = act_a.split(" (")[1].rstrip(")") if "(" in act_a else ""
            gv_b = act_b.split(" (")[1].rstrip(")") if "(" in act_b else ""
            
            grid_a = teacher_grids.get(gv_a)
            grid_b = teacher_grids.get(gv_b)
            if not grid_a or not grid_b: continue
            
            # Can gv_a and gv_b go to any distinct slots in dest_slots?
            for d_a in dest_slots:
                for d_b in dest_slots:
                    if d_a == d_b: continue
                    if grid_a[d_a] == "" and grid_b[d_b] == "":
                        print(f" -> MATCH in {sess_name}: Move 7A17-Toán to slots [{s_a}, {s_b}]; Move {act_a} to slot {d_a}; Move {act_b} to slot {d_b}!")

