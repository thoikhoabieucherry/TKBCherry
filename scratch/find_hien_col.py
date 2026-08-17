import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r"C:\Users\Love\Documents\Codex\TKBCherry\temp\tonggv0917082026.xlsx", data_only=True)
ws = wb.active

for col in range(1, ws.max_column + 1):
    vals = [ws.cell(r, col).value for r in range(1, 10)]
    for r, v in enumerate(vals, 1):
        if v and ("Hiền" in str(v) or "hien" in str(v).lower()):
            print(f"Col {col} (Row {r}): {v}")
            # Print all scheduled cells in this column
            schedule = []
            for r2 in range(1, ws.max_row + 1):
                cell_val = ws.cell(r2, col).value
                day_val = ws.cell(r2, 1).value
                session_val = ws.cell(r2, 2).value
                period_val = ws.cell(r2, 3).value
                if cell_val:
                    schedule.append((r2, day_val, session_val, period_val, cell_val))
            print(f"  Schedule for col {col}:")
            for item in schedule:
                print(f"    Row {item[0]}: {item[1]} {item[2]} T{item[3]} -> {item[4]}")
