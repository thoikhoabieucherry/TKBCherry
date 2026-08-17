import openpyxl
import os
import sys
import json
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r"C:\Users\Love\Documents\Codex\temp\xepmoi.xlsx", data_only=True)
sh = wb["TKB_LOP_SC"]

classes_set = []
teachers_set = set()
subjects_set = set()
pccmMatrix = {}
pccmTietMatrix = defaultdict(int)

col_to_class_session = {}
for c in range(3, sh.max_column + 1):
    c_name = sh.cell(4, c).value
    if c_name:
        c_str = str(c_name).strip()
        if c_str not in classes_set:
            classes_set.append(c_str)
        sess_str = str(sh.cell(5, c).value or "").strip().lower()
        sess = "sang" if "sáng" in sess_str or "sang" in sess_str else "chieu"
        col_to_class_session[c] = (c_str, sess)

DAYS_KEYS = ["thu2", "thu3", "thu4", "thu5", "thu6", "thu7"]

tkb_data = defaultdict(lambda: {
    d: { "sang": ["OFF"] * 5, "chieu": ["OFF"] * 5 } for d in DAYS_KEYS
})

for r in range(6, 36):
    r_idx = r - 6
    d_idx = r_idx // 5
    p_idx = r_idx % 5
    day_key = DAYS_KEYS[d_idx]
    
    for c, (c_name, sess) in col_to_class_session.items():
        cell_val = str(sh.cell(r, c).value or "").strip()
        if cell_val and cell_val != "-" and cell_val != "OFF":
            parts = cell_val.split("-")
            mon = parts[0].strip()
            gv = parts[1].strip() if len(parts) > 1 else ""
            
            subjects_set.add(mon)
            if gv:
                teachers_set.add(gv)
                pccmMatrix[f"{c_name}|{mon}"] = gv
                pccmTietMatrix[f"{c_name}|{mon}"] += 1
                
            tkb_data[c_name][day_key][sess][p_idx] = mon
        else:
            tkb_data[c_name][day_key][sess][p_idx] = "OFF"

lop_list = [{ "id": c, "ten": c, "ten2": c } for c in classes_set]
giaovien_list = [{ "id": gv, "ten": gv, "magv": gv } for gv in sorted(list(teachers_set))]
monhoc_list = [{ "id": m, "ten": m, "ten2": m } for m in sorted(list(subjects_set))]

school_obj = {
    "lop": lop_list,
    "giaovien": giaovien_list,
    "monhoc": monhoc_list,
    "pccmMatrix": pccmMatrix,
    "pccmTietMatrix": dict(pccmTietMatrix),
    "tkb": { c: days for c, days in tkb_data.items() },
    "tkbConstraints": {}
}

with open(r"C:\Users\Love\Documents\Codex\TKBCherry\scratch\dongkhoi_base_vps.json", "w", encoding="utf-8") as f:
    json.dump(school_obj, f, ensure_ascii=False, indent=2)

print(f"Created dongkhoi_base_vps.json: {len(lop_list)} classes, {len(giaovien_list)} teachers, {len(monhoc_list)} subjects, {len(pccmMatrix)} PCCM assignments")
