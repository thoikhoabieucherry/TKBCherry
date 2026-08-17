import sys
import openpyxl
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

file_path = Path(r"C:\Users\Love\Documents\Codex\TKBCherry\temp\tonggv0917082026.xlsx")
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb.active

# Let's inspect the schedule of 7A17 and 7A18 across all teachers
# For each time slot (row 3 to 62), find what class is assigned across all teachers
def get_class_schedule(class_name):
    print("\n" + "="*70)
    print(f"THỜI KHÓA BIỂU TOÀN BỘ CỦA LỚP: {class_name}")
    print("="*70)
    for r in range(3, 63):
        d_val = sheet.cell(r, 1).value
        b_val = sheet.cell(r, 2).value
        p_val = sheet.cell(r, 3).value
        
        # search across all teacher columns
        lessons = []
        for c in range(4, sheet.max_column + 1):
            v = sheet.cell(r, c).value
            if v and class_name in str(v):
                gv_name = sheet.cell(2, c).value
                lessons.append(f"{str(v).strip()} (GV: {gv_name})")
        
        lesson_str = ", ".join(lessons) if lessons else "[TRỐNG]"
        print(f"  Thứ {d_val if d_val else ' '} - {b_val if b_val else ' ':<5} - Tiết {p_val}: {lesson_str}")

get_class_schedule("7A17")
get_class_schedule("7A18")

