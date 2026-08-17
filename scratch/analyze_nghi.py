import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook("C:/Users/Love/Documents/Codex/phancong/tesst/nghi.xlsx", data_only=True)
print("Sheets in nghi.xlsx:", wb.sheetnames)

for sname in wb.sheetnames:
    sheet = wb[sname]
    print(f"\n--- Sheet: {sname} ---")
    for r in range(1, min(35, sheet.max_row + 1)):
        row_vals = [str(sheet.cell(r, c).value or "").strip() for c in range(1, min(25, sheet.max_column + 1))]
        if any(row_vals):
            print(f"Row {r:2d}: {row_vals}")
